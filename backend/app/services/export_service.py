"""
PowerX 数据导出服务

创建日期: 2026-01-07
作者: zhi.qu

提供数据导出功能：CSV、Excel、PDF
"""

import io
import csv
import json
from datetime import datetime, date
from typing import Any, Dict, List, Optional, Union
from loguru import logger


class ExportService:
    """
    数据导出服务
    
    支持格式：
    - CSV
    - Excel (需要 openpyxl)
    - JSON
    - PDF (需要 reportlab)
    """
    
    def __init__(self):
        logger.info("导出服务初始化")
    
    def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None,
        headers: Optional[Dict[str, str]] = None
    ) -> bytes:
        """
        导出为 CSV 格式
        
        Args:
            data: 数据列表
            columns: 要导出的列
            headers: 列标题映射 {字段名: 显示名}
            
        Returns:
            bytes: CSV 文件内容
        """
        if not data:
            return b""
        
        # 确定列
        if columns is None:
            columns = list(data[0].keys())
        
        # 确定标题
        if headers is None:
            headers = {col: col for col in columns}
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 写入标题行
        writer.writerow([headers.get(col, col) for col in columns])
        
        # 写入数据行
        for row in data:
            writer.writerow([self._format_value(row.get(col, "")) for col in columns])
        
        content = output.getvalue()
        output.close()
        
        # 添加 BOM 以支持 Excel 中的中文
        return b'\xef\xbb\xbf' + content.encode('utf-8')
    
    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None,
        headers: Optional[Dict[str, str]] = None,
        sheet_name: str = "Sheet1"
    ) -> bytes:
        """
        导出为 Excel 格式
        
        Args:
            data: 数据列表
            columns: 要导出的列
            headers: 列标题映射
            sheet_name: 工作表名称
            
        Returns:
            bytes: Excel 文件内容
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            logger.warning("openpyxl 未安装，回退到 CSV 格式")
            return self.export_to_csv(data, columns, headers)
        
        if not data:
            return b""
        
        # 确定列
        if columns is None:
            columns = list(data[0].keys())
        
        # 确定标题
        if headers is None:
            headers = {col: col for col in columns}
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入标题行
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=headers.get(col_name, col_name))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 写入数据行
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = self._format_value(row_data.get(col_name, ""))
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        # 自动调整列宽
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(str(headers.get(col_name, col_name)))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.read()
    
    def export_to_json(
        self,
        data: List[Dict[str, Any]],
        indent: int = 2
    ) -> bytes:
        """
        导出为 JSON 格式
        
        Args:
            data: 数据列表
            indent: 缩进空格数
            
        Returns:
            bytes: JSON 文件内容
        """
        return json.dumps(data, ensure_ascii=False, indent=indent, default=str).encode('utf-8')
    
    def _format_value(self, value: Any) -> str:
        """格式化单元格值"""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, bool):
            return "是" if value else "否"
        if isinstance(value, (list, dict)):
            return json.dumps(value, ensure_ascii=False)
        return str(value)
    
    # ============ 预定义导出模板 ============
    
    def export_trading_orders(self, orders: List[Dict]) -> bytes:
        """导出交易订单"""
        columns = ["order_id", "order_type", "direction", "quantity", "price", "status", "created_at"]
        headers = {
            "order_id": "订单号",
            "order_type": "订单类型",
            "direction": "方向",
            "quantity": "数量(MWh)",
            "price": "价格(元/MWh)",
            "status": "状态",
            "created_at": "创建时间"
        }
        return self.export_to_excel(orders, columns, headers, "交易订单")
    
    def export_contracts(self, contracts: List[Dict]) -> bytes:
        """导出合同数据"""
        columns = ["contract_id", "contract_type", "counterparty", "total_quantity", "price", "start_date", "end_date", "status"]
        headers = {
            "contract_id": "合同编号",
            "contract_type": "合同类型",
            "counterparty": "交易对手",
            "total_quantity": "总电量(MWh)",
            "price": "单价(元/MWh)",
            "start_date": "开始日期",
            "end_date": "结束日期",
            "status": "状态"
        }
        return self.export_to_excel(contracts, columns, headers, "合同列表")
    
    def export_settlement_records(self, records: List[Dict]) -> bytes:
        """导出结算记录"""
        columns = ["settlement_id", "period", "electricity_fee", "capacity_fee", "ancillary_fee", "total_amount", "status"]
        headers = {
            "settlement_id": "结算单号",
            "period": "结算周期",
            "electricity_fee": "电量电费",
            "capacity_fee": "容量电费",
            "ancillary_fee": "辅助服务费",
            "total_amount": "总金额",
            "status": "状态"
        }
        return self.export_to_excel(records, columns, headers, "结算记录")
    
    def export_market_prices(self, prices: List[Dict]) -> bytes:
        """导出市场价格"""
        columns = ["date", "hour", "province", "day_ahead_price", "realtime_price", "volume"]
        headers = {
            "date": "日期",
            "hour": "时段",
            "province": "省份",
            "day_ahead_price": "日前价格",
            "realtime_price": "实时价格",
            "volume": "成交量"
        }
        return self.export_to_excel(prices, columns, headers, "市场价格")


# 全局导出服务实例
export_service = ExportService()
