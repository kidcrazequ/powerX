"""
PowerX 数据导入服务

创建日期: 2026-01-07
作者: zhi.qu

提供数据导入功能：CSV、Excel、JSON
"""

import io
import csv
import json
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple, Callable
from loguru import logger


class ImportError(Exception):
    """导入错误"""
    def __init__(self, message: str, row: int = 0, errors: List[Dict] = None):
        super().__init__(message)
        self.row = row
        self.errors = errors or []


class ImportService:
    """
    数据导入服务
    
    支持格式：
    - CSV
    - Excel (需要 openpyxl)
    - JSON
    """
    
    def __init__(self):
        logger.info("导入服务初始化")
    
    def parse_csv(
        self,
        file_content: bytes,
        column_mapping: Optional[Dict[str, str]] = None,
        validators: Optional[Dict[str, Callable]] = None
    ) -> Tuple[List[Dict], List[Dict]]:
        """
        解析 CSV 文件
        
        Args:
            file_content: 文件内容
            column_mapping: 列名映射 {CSV列名: 字段名}
            validators: 字段验证器 {字段名: 验证函数}
            
        Returns:
            Tuple: (成功数据列表, 错误列表)
        """
        # 处理 BOM
        if file_content.startswith(b'\xef\xbb\xbf'):
            file_content = file_content[3:]
        
        # 尝试不同编码
        for encoding in ['utf-8', 'gbk', 'gb2312']:
            try:
                content = file_content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ImportError("无法识别文件编码")
        
        reader = csv.DictReader(io.StringIO(content))
        
        success_data = []
        errors = []
        
        for row_idx, row in enumerate(reader, 2):  # 从第2行开始（第1行是标题）
            try:
                # 映射列名
                if column_mapping:
                    mapped_row = {}
                    for csv_col, field_name in column_mapping.items():
                        if csv_col in row:
                            mapped_row[field_name] = row[csv_col]
                    row = mapped_row
                
                # 验证数据
                if validators:
                    for field, validator in validators.items():
                        if field in row:
                            row[field] = validator(row[field])
                
                success_data.append(row)
                
            except Exception as e:
                errors.append({
                    "row": row_idx,
                    "data": dict(row),
                    "error": str(e)
                })
        
        logger.info(f"CSV解析完成: 成功 {len(success_data)} 条, 失败 {len(errors)} 条")
        
        return success_data, errors
    
    def parse_excel(
        self,
        file_content: bytes,
        sheet_name: Optional[str] = None,
        column_mapping: Optional[Dict[str, str]] = None,
        validators: Optional[Dict[str, Callable]] = None
    ) -> Tuple[List[Dict], List[Dict]]:
        """
        解析 Excel 文件
        
        Args:
            file_content: 文件内容
            sheet_name: 工作表名称
            column_mapping: 列名映射
            validators: 字段验证器
            
        Returns:
            Tuple: (成功数据列表, 错误列表)
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl 未安装，无法解析 Excel 文件")
        
        wb = load_workbook(io.BytesIO(file_content), read_only=True)
        
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows:
            return [], []
        
        # 第一行作为标题
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
        
        success_data = []
        errors = []
        
        for row_idx, row in enumerate(rows[1:], 2):
            try:
                row_dict = dict(zip(headers, row))
                
                # 映射列名
                if column_mapping:
                    mapped_row = {}
                    for excel_col, field_name in column_mapping.items():
                        if excel_col in row_dict:
                            mapped_row[field_name] = row_dict[excel_col]
                    row_dict = mapped_row
                
                # 验证数据
                if validators:
                    for field, validator in validators.items():
                        if field in row_dict and row_dict[field] is not None:
                            row_dict[field] = validator(row_dict[field])
                
                success_data.append(row_dict)
                
            except Exception as e:
                errors.append({
                    "row": row_idx,
                    "data": dict(zip(headers, row)),
                    "error": str(e)
                })
        
        wb.close()
        
        logger.info(f"Excel解析完成: 成功 {len(success_data)} 条, 失败 {len(errors)} 条")
        
        return success_data, errors
    
    def parse_json(
        self,
        file_content: bytes,
        validators: Optional[Dict[str, Callable]] = None
    ) -> Tuple[List[Dict], List[Dict]]:
        """
        解析 JSON 文件
        
        Args:
            file_content: 文件内容
            validators: 字段验证器
            
        Returns:
            Tuple: (成功数据列表, 错误列表)
        """
        try:
            data = json.loads(file_content.decode('utf-8'))
        except json.JSONDecodeError as e:
            raise ImportError(f"JSON 解析错误: {str(e)}")
        
        if not isinstance(data, list):
            data = [data]
        
        success_data = []
        errors = []
        
        for row_idx, row in enumerate(data, 1):
            try:
                if not isinstance(row, dict):
                    raise ValueError("数据格式错误，期望字典类型")
                
                # 验证数据
                if validators:
                    for field, validator in validators.items():
                        if field in row:
                            row[field] = validator(row[field])
                
                success_data.append(row)
                
            except Exception as e:
                errors.append({
                    "row": row_idx,
                    "data": row,
                    "error": str(e)
                })
        
        logger.info(f"JSON解析完成: 成功 {len(success_data)} 条, 失败 {len(errors)} 条")
        
        return success_data, errors
    
    def auto_parse(
        self,
        file_content: bytes,
        filename: str,
        column_mapping: Optional[Dict[str, str]] = None,
        validators: Optional[Dict[str, Callable]] = None
    ) -> Tuple[List[Dict], List[Dict]]:
        """
        根据文件扩展名自动选择解析器
        
        Args:
            file_content: 文件内容
            filename: 文件名
            column_mapping: 列名映射
            validators: 字段验证器
            
        Returns:
            Tuple: (成功数据列表, 错误列表)
        """
        ext = filename.lower().split('.')[-1]
        
        if ext == 'csv':
            return self.parse_csv(file_content, column_mapping, validators)
        elif ext in ['xlsx', 'xls']:
            return self.parse_excel(file_content, None, column_mapping, validators)
        elif ext == 'json':
            return self.parse_json(file_content, validators)
        else:
            raise ImportError(f"不支持的文件格式: {ext}")
    
    # ============ 预定义验证器 ============
    
    @staticmethod
    def validate_float(value: Any) -> float:
        """验证并转换为浮点数"""
        if value is None or value == "":
            return 0.0
        return float(value)
    
    @staticmethod
    def validate_int(value: Any) -> int:
        """验证并转换为整数"""
        if value is None or value == "":
            return 0
        return int(float(value))
    
    @staticmethod
    def validate_date(value: Any) -> str:
        """验证并转换日期"""
        if value is None or value == "":
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        # 尝试解析常见日期格式
        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"]:
            try:
                return datetime.strptime(str(value), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return str(value)
    
    @staticmethod
    def validate_required(value: Any) -> Any:
        """验证必填字段"""
        if value is None or value == "":
            raise ValueError("此字段为必填项")
        return value
    
    # ============ 预定义导入模板 ============
    
    def import_trading_orders(self, file_content: bytes, filename: str) -> Tuple[List[Dict], List[Dict]]:
        """导入交易订单"""
        column_mapping = {
            "订单类型": "order_type",
            "方向": "direction",
            "数量(MWh)": "quantity",
            "价格(元/MWh)": "price",
            "省份": "province"
        }
        validators = {
            "quantity": self.validate_float,
            "price": self.validate_float
        }
        return self.auto_parse(file_content, filename, column_mapping, validators)
    
    def import_contracts(self, file_content: bytes, filename: str) -> Tuple[List[Dict], List[Dict]]:
        """导入合同数据"""
        column_mapping = {
            "合同类型": "contract_type",
            "交易对手": "counterparty",
            "总电量(MWh)": "total_quantity",
            "单价(元/MWh)": "price",
            "开始日期": "start_date",
            "结束日期": "end_date"
        }
        validators = {
            "total_quantity": self.validate_float,
            "price": self.validate_float,
            "start_date": self.validate_date,
            "end_date": self.validate_date
        }
        return self.auto_parse(file_content, filename, column_mapping, validators)


# 全局导入服务实例
import_service = ImportService()
